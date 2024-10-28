from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common import TimeoutException
from pyquery import PyQuery as pq
from urllib.parse import quote
import time
from lxml import etree
import requests
import openpyxl
import ast
import xlsxwriter
import sys

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "localhost:9222") #此处端口保持和命令行启动的端口一致
driver = Chrome(options=chrome_options)
actions = ActionChains(driver)
wait = WebDriverWait(driver, 10)


# chrome.exe --remote-debugging-port=9222 --user-data-dir='D:\chrome_data'
def get_item(url, paipai):
    result = {}  # 手机参数
    comment_list = []
    hp_percent = 0
    try:
        driver.get(url)
        time.sleep(2)
        html = driver.page_source
        doc = pq(html)
        try:
            parameter_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="detail"]/div[1]/ul/li[2]')))  # 参数按钮
            parameter_btn.click()
            time.sleep(1.5)
        except TimeoutException:
            return str(result), comment_list, str(hp_percent)
        items = doc('.clearfix').items()  # .Ptable-item
        for item in items:
            key = item('dt').text()
            content = item('dd').text()
            # key = item.find('.clearfix')('dt').text()
            # content = item.find('.clearfix')('dd').text()
            if key == '' or item == '':
                continue
            if key != '入网型号':
                result[key] = content
            else:
                if len(content) > 10:
                    result[key] = content[11:]
        print(result)
        if paipai == 1:  #
            try:
                comment_btn = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="detail"]/div[1]/ul/li[4]')))  # 参数按钮
                comment_btn.click()
            except TimeoutException:
                return str(result), comment_list, str(hp_percent)
        else:
            try:
                comment_btn = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="detail"]/div[1]/ul/li[5]')))  # 参数按钮
                comment_btn.click()
            except TimeoutException:
                return str(result), comment_list, str(hp_percent)
        time.sleep(0.5)
        html = driver.page_source
        doc = pq(html)
        hp_percent = doc('.percent-con').text()
        print('好评率', hp_percent)

        # 差评
        try:
            bad_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="comment"]/div[2]/div[2]/div[1]/ul/li[7]/a')))  # 参数按钮
            bad_btn.click()
        except TimeoutException:
            return str(result), comment_list, str(hp_percent)
        time.sleep(1.5)
        html = driver.page_source
        doc = pq(html)
        comments = doc('.comment-con').items()  # ('.J-comments-list comments-list ETab')('.tab-con')
        for item in comments:
            comment_list.append('。'.join(item.text().split('\n')))

        # 中评
        try:
            mid_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="comment"]/div[2]/div[2]/div[1]/ul/li[6]/a')))  # 参数按钮
            mid_btn.click()
        except TimeoutException:
            return str(result), comment_list, str(hp_percent)
        time.sleep(1.5)
        html = driver.page_source
        doc = pq(html)
        comments = doc('.comment-con').items()  # ('.J-comments-list comments-list ETab')('.tab-con')
        for item in comments:
            comment_list.append('。'.join(item.text().split('\n')))

        # 好评
        try:
            good_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="comment"]/div[2]/div[2]/div[1]/ul/li[1]/a')))  # 参数按钮
            good_btn.click()
        except TimeoutException:
            return str(result), comment_list, str(hp_percent)
        time.sleep(0.5)
        max_page = 2
        for i in range(max_page):  # 评论翻页
            if i != 0:
                html = driver.page_source
                doc = pq(html)
                if doc.find('.ui-pager-next').text():  # 下一页
                    try:
                        next_page_btn = driver.find_element(By.XPATH, '//*[@id="comment-0"]/div[12]/div/div/a[2]')
                        driver.execute_script('arguments[0].click();', next_page_btn)
                    except TimeoutException:
                        return str(result), comment_list, str(hp_percent)
                else:
                    break
            time.sleep(4)
            html = driver.page_source
            doc = pq(html)
            comments = doc('.comment-con').items()  # ('.J-comments-list comments-list ETab')('.tab-con')
            for item in comments:
                comment_list.append('。'.join(item.text().split('\n')))

        print(str(comment_list))
    except TimeoutException:
        return str(result), comment_list, str(hp_percent)
    return str(result), comment_list, str(hp_percent)


def get_jd_item():
    wb = openpyxl.load_workbook(filename="jd.xlsx")
    ws = wb['Sheet1']
    # count = 1
    count = 433
    continue_x = True
    workbook = xlsxwriter.Workbook('jd_item.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, 'index')
    worksheet.write(0, 1, 'id')
    worksheet.write(0, 2, 'name')
    worksheet.write(0, 3, 'price')
    worksheet.write(0, 4, 'category')
    worksheet.write(0, 5, 'picture_addr')
    worksheet.write(0, 6, 'specifications')
    worksheet.write(0, 7, 'Feedback_Rate')
    worksheet.write(0, 8, 'comment')
    i = 1
    wb_help = openpyxl.load_workbook(filename="jd_item.xlsx")
    ws_help = wb_help['Sheet1']
    for row in ws_help.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, values_only=True):
        if i < count and continue_x:
            worksheet.write(i, 0, row[0])
            worksheet.write(i, 1, row[1])
            worksheet.write(i, 2, row[2])
            worksheet.write(i, 3, row[3])
            worksheet.write(i, 4, row[4])
            worksheet.write(i, 5, row[5])
            worksheet.write(i, 6, row[6])
            worksheet.write(i, 7, row[7])
            worksheet.write(i, 8, row[8])
            i += 1
        else:

            break
    i = 1
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, values_only=True):  # ws.max_row
        if i < count and continue_x:
            i += 1
            continue
        continue_x = False
        id = row[1]
        url = "https://item.jd.com/" + str(id) + ".html"
        worksheet.write(count, 0, row[0])
        worksheet.write(count, 1, url)
        worksheet.write(count, 2, row[2])
        name = row[2]
        paipai = 0
        if name[:2] == '拍拍':
            paipai = 1
        spec, comments, hp_percent = get_item(url, paipai)
        worksheet.write(count, 3, row[3])
        worksheet.write(count, 4, row[4])
        worksheet.write(count, 5, row[5])
        worksheet.write(count, 7, hp_percent)
        if len(spec) == 0:
            time.sleep(10)
            continue
        else:
            worksheet.write(count, 6, spec)
        if len(comments):  # comments不为空
            # for index, comment in enumerate(comments):
            #     print(comment)
            #     worksheet.write(count, index + 8, comment)
            worksheet.write(count, 8, str(comments))
        elif len(comments) == 0 and hp_percent=='': #  and hp_percent == 0
            print('empty')
            print(hp_percent)
            print(count)
            break
        count += 1
    workbook.close()


def read_only():
    wb = openpyxl.load_workbook(filename="jd_item.xlsx")
    ws = wb['Sheet1']
    lis = []
    dic = {"id": None, "name": None, "price": None, "category": None, 'picture_addr': None, 'specification': None,
           'comments': None, 'Feedback_Rate': None}
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, values_only=True):
        # print(row)
        dic["id"] = row[0]
        dic["name"] = row[1]
        dic["price"] = row[2]
        dic["category"] = row[4]
        dic["picture_addr"] = row[5]
        c = ast.literal_eval(row[6])
        # print(c, type(c))
        dic["specification"] = c
        dic['Feedback_Rate'] = row[7]
        dic['comments'] = row[8]
        lis.append(dic)
    # print(lis)
    return lis


if __name__ == '__main__':
    get_jd_item()  # 爬取手机参数和评价
    # read_only()
