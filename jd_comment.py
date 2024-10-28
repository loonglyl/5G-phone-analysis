import requests
import json
import time
import random
import xlwt
import xlutils.copy
import xlrd
import openpyxl

def start(page, score, id):
    # 获取URL
    #score 评价等级 page=0 第一页 producitid 商品类别
    url = 'https://club.jd.com/comment/productPageComments.action?&productId=' + str(id) + '&score=' + str(score) + '&sortType=5&page='+ str(page) + '&pageSize=10&isShadowSku=0&fold=1'
    headers= {
        "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Mobile Safari/537.36"
    }
    time.sleep(2)
    test = requests.get(url=url, headers= headers)
    data = json.loads(test.text)
    return data
    # 解析页面
def parse(data):

    items = data['comments']
    for i in items:
        yield (
            i['nickname'],#用户名字
            i['id'], #用户id
            i['content'],#内容
            i['creationTime']#时间
        )

def excel(items, id):
    #第一次写入
    newTable = str(id) + ".xlsx"#创建文件
    wb = xlwt.Workbook("encoding='utf-8")

    ws = wb.add_sheet('sheet1')#创建表
    headDate = ['用户名','id', '内容', '时间']#定义标题
    for i in range(0,4):#for循环遍历写入
        ws.write(0, i, headDate[i], xlwt.easyxf('font: bold on'))

    index = 1#行数

    for data in items:#items是十条数据 data是其中一条（一条下有三个内容）
        for i in range(0,4):#列数

            print(data[i])
            ws.write(index, i, data[i])#行 列 数据（一条一条自己写入）
        print('______________________')
        index += 1#等上一行写完了 在继续追加行数
        wb.save(newTable)

def another(items, j, id):#如果不是第一次写入 以后的就是追加数据了 需要另一个函数

    index = (j-1) * 10 + 1#这里是 每次写入都从11 21 31..等开始 所以我才传入数据 代表着从哪里开始写入

    data = xlrd.open_workbook(str(id) + ".xlsx")
    ws = xlutils.copy.copy(data)
    # 进入表
    table = ws.get_sheet(0)

    for test in items:

        for i in range(0, 4):#跟excel同理
            print(test[i])

            table.write(index, i, test[i])  # 只要分配好 自己塞入
        print('_______________________')

        index += 1
        ws.save(str(id) + ".xlsx")



def main():
    j = 1#页面数
    judge = True#判断写入是否为第一次写入

    wb = openpyxl.load_workbook(filename="jd.xlsx")
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=151, values_only=True):
        print(row[0])
        for k in range(0, 2):
            for i in range(0, 10):
                time.sleep(1.5)
                # 记得time反爬 其实我在爬取的时候没有使用代理ip也没给我封 不过就当这是个习惯吧
                first = start(j, k, row[0])
                test = parse(first)
                if judge:
                    excel(test, row[0])
                    judge = False
                else:
                    another(test, j, row[0])
                print('第' + str(j) + '页抓取完毕\n')
                j = j + 1
        break


if __name__ == '__main__':
    main()
    # 这个代码仅为全部数据下的评论而已 中差评等需要修改score！