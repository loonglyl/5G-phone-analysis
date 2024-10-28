from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common import TimeoutException
from pyquery import PyQuery as pq
from urllib.parse import quote
import time
import openpyxl
import copy

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")  # 此处端口保持和命令行启动的端口一致
driver = Chrome(options=chrome_options, )  # executable_path='D:\Anaconda3\chrome-win64\chromedriver.exe'
wait = WebDriverWait(driver, 10)


wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0)
ws.cell(row=1, column=1, value="index")
ws.cell(row=1, column=2, value="id")
ws.cell(row=1, column=3, value="name")
ws.cell(row=1, column=4, value="price")
ws.cell(row=1, column=5, value="category")
ws.cell(row=1, column=6, value="picture_addr")


def index_page(keyword, max_page, count):
    search = keyword + '手机'
    url = "https://search.jd.com/Search?keyword=" + search
    # btn = wait.until(EC.element_to_be_clickable(
    #     (By.XPATH, '//*[@id="J_selector"]/div[1]/div/div[2]/div[1]/ul/li[1]/a')))  # 参数按钮
    # btn.click()
    driver.get(url)
    time.sleep(1.5)

    # 获取商品图片
    for i in range(max_page):
        driver.refresh()
        html = driver.page_source
        doc = pq(html)
        time.sleep(0.5)
        for y in range(28):
            js = 'window.scrollBy(0,200)'
            driver.execute_script(js)
            time.sleep(0.5)
        html = driver.page_source
        doc = pq(html)
        items = doc('.gl-i-wrap').items()  # 所有手机商品
        for item in items:
            # 获取商品信息
            count += 1
            product = {'id': item.find('.p-icons').attr('id')[6:],
                       'price': item.find('.p-price')('i').text(),
                       'title': ''.join(item.find('.p-name')('a')('em').text().split('\n'))}
            ws.cell(row=count, column=1, value=str(count - 1))
            ws.cell(row=count, column=2, value=str(product['id']))
            ws.cell(row=count, column=3, value=str(product['title']))
            ws.cell(row=count, column=4, value=str(product['price']))
            print(product)

            print(item.find('.p-img')('img').attr('src'))
            picture = {'category': keyword,
                       'picture_addr': item.find('.p-img')('img').attr('src')}
            if picture['picture_addr'] != "None":
                picture['picture_addr'] = "http:" + picture['picture_addr']
            ws.cell(row=count, column=5, value=str(picture['category']))
            ws.cell(row=count, column=6, value=str(picture['picture_addr']))
            print(picture)
        if i != max_page-1:
            next_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="J_bottomPage"]/span[1]/a[9]')))  # 参数按钮
            next_btn.click()
            time.sleep(1.5)
    return count


def read_only():
    wb = openpyxl.load_workbook(filename="jd.xlsx")
    ws = wb['Sheet1']
    lis = []
    dic = {"index": None, "id": None, "name": None, "price": None, "category": None, "picture_addr": None}
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, values_only=True):  # , max_col=5
        # print(row)
        dic["index"] = row[0]
        dic['id'] = row[1]
        dic['name'] = row[2]
        dic['price'] = row[3]
        dic["category"] = row[4]
        dic["picture_addr"] = row[5]
        lis.append(copy.deepcopy(dic))
    print(lis)
    return lis


if __name__ == '__main__':
    # category = ['华为', '小米', 'apple', '荣耀', '三星', '红米', 'oppo', 'vivo', '一加', '魅族', 'iqoo', '真我']
    # count = 1
    # for i in category:
    #     keyword = i
    #     max_page = 1
    #     count = index_page(keyword, max_page, count)
    # wb.save("jd.xlsx")  # 保存手机商品的网址信息
    read_only()
