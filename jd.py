from lxml import etree
import requests
import time
import openpyxl

outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)

outws.cell(row=1, column=1, value="index")
outws.cell(row=1, column=2, value="id")
outws.cell(row=1, column=3, value="name")
outws.cell(row=1, column=4, value="price")


def jd_page():
    page = 1
    s = 1
    count = 1
    for i in range(1, 3):
        print("page=" + str(page) + ",s=" + str(s))
        url = "https://search.jd.com/Search?keyword=5G手机&wq=5G手机&pvid=feb1e81188ca4a11964991bce3bbf228&isList=0&page="+str(page)+"&s="+str(s)+"&click=0"
        page = page+1
        s = s+30
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0",
            "Cookie": "shshshfpa=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; shshshfpx=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; __jdu=1712498482836958502465; pinId=BKmhqjK0Ot4FvWOdtn9TiA; pin=jd_lTmXtaLBvUsY; unick=jd_lTmXtaLBvUsY; _tp=7%2B9vkUgexJVrllJMczJ5lw%3D%3D; _pst=jd_lTmXtaLBvUsY; user-key=d4e0b5c4-efd8-4e69-87e6-a0f85f57fe3a; qrsc=3; unpl=JF8EALFnNSttWk1dAxsHThtHQ1oHW1gATEQBamQFAAlcTl1QSAodQUV7XlVdWBRKEh9vYRRVVVNKXA4aACsSEXteU11bD00VB2xXXAQDGhUQR09SWEBJJVlQXl4ITxcFZ2A1ZF5Ye1QEKwIdEhNOVV1bVAtCFQRnYwxQXFtOUwErAysVIHttU1ZYDUoQM25XBGQfDBdWAR0HGRtdS1tUXVgAQhIKbG4HU1VcQlAEGAccFiBKbVc; __jdv=76161171|haosou-search|t_262767352_haosousearch|cpc|5512151796_0_378714d8f96b4496b3421ee548db97bd|1713316144511; areaId=19; ipLoc-djd=19-1601-50258-129167; TrackID=1JSb7ijeU0EfdV5-iYXc4__5LdRAVrkfc97Iias3TkhIZsyRi6cA5FBrkNtYGRbP1WtiIvD7yahJxauK6S0JYVG9it7lQFlrvRA4QzTFYeSfUyEO84P7E9vWFMUezQeny; thor=676741B8CA3FCE6E194F2D84F06C5CB3344018AEFA0977C08C33E05EF2AEFFC2D0A42EA46D88CF2D6B178FDAAA985F63FAA583759B36AB62234AB1B33A45A455ACF514C7394AA90420946DB02EBFC9C9086C3CF93A55E3FBDCCB979E1BD583B396581BA939E938169295926E7780CF82164F7B763197BB10E8D1B3CC51D2C0BD2D5583BD01464691741F806AECA53B8CB703008F72C1F3F11C1FDB7DC4C60A0A; flash=2_Ts9f6zFn2Tqw0NqMxG4hqLTQ13GCG1ID25q2whi2C-wINYjIm1mJlK882w-DJEDo53i-11Mh_olckTVhMFBu14BGBMXX3jll3hQvglus2d5*; jsavif=1; jsavif=1; _gia_d=1; 3AB9D23F7A4B3CSS=jdd032ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQAAAAMPBW7FEQAAAAAACFT2A57E3VCIIUX; shshshfpb=BApXckia2DupAmC3SbfNMBEgpx2ei76hCBlAIg6169xJ1MrE1XoO2; avif=1; xapieid=jdd032ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQAAAAMPBW7FEQAAAAAACFT2A57E3VCIIUX; __jda=143920055.1712498482836958502465.1712498483.1713749255.1713922086.10; __jdb=143920055.5.1712498482836958502465|10.1713922086; __jdc=143920055; rkv=1.0; 3AB9D23F7A4B3C9B=2ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQ"
        }
        res = requests.get(url, headers=headers)
        time.sleep(1.5)
        res.encoding = 'utf-8'
        text = res.text

        selector = etree.HTML(text)
        list = selector.xpath('//*[@id="J_goodsList"]/ul/li')

        for i in list:
            product_name = i.xpath('.//div[@class="p-name p-name-type-2"]/a/em/text()')[0]
            product_price = i.xpath('.//div[@class="p-price"]/strong/i/text()')[0]
            product_id = i.xpath('.//div[@class="p-commit"]/strong/a/@id')[0].replace("J_comment_", "")
            print("product_id= " + str(product_id))
            print("product_name= " + str(product_name))
            print("product_price= " + str(product_price))
            print("-----")
            count += 1
            outws.cell(row=count, column=1, value=str(count - 1))
            outws.cell(row=count, column=2, value=str(product_id))
            outws.cell(row=count, column=3, value=str(product_name))
            outws.cell(row=count, column=4, value=str(product_price))


jd_page()
outwb.save("jd.xlsx")  #保存