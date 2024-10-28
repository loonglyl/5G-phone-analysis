from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common import TimeoutException
from pyquery import PyQuery as pq
from urllib.parse import quote
import time
from lxml import etree
import requests
import openpyxl
import ast
import xlsxwriter

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "localhost:9222") #此处端口保持和命令行启动的端口一致
driver = Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)


# 模拟淘宝登录
# chrome.exe --remote-debugging-port=9222 --user-data-dir='D:\chrome_data'
def login_taobao():
    print('开始登录...')
    login_url = 'https://login.taobao.com/member/login.jhtml'
    try:
        driver.get(login_url)
        input_login_id = wait.until(EC.presence_of_element_located((By.ID, 'fm-login-id')))
        input_login_password = wait.until(EC.presence_of_element_located((By.ID, 'fm-login-password')))
        input_login_id.send_keys('')  # 用你自己的淘宝账号替换
        input_login_password.send_keys('Alihualin123')  # 用你自己的密码替换
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.fm-button.fm-submit.password-login')))
        submit.click()
        is_loging = wait.until(EC.url_changes(login_url))
        return is_loging
    except TimeoutException:
        print('login_taobao TimeoutException')
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.fm-button.fm-submit')))
        submit.click()
        is_loging = wait.until(EC.url_changes(login_url))
        if is_loging:
            return is_loging
        else:
            login_taobao()


def get_item_comment(url):
    if url[0:6] != 'https:':
        url = 'https:' + url
    # print(url)
    error = False
    result = {}  # 手机参数
    comments = []  # 手机评论
    try:
        driver.get(url)
        time.sleep(0.5)
        html = driver.page_source
        doc = pq(html)
        # items = doc('.ItemDetail--attrs--3t-mTb3').items()
        # items = doc('.Attrs--attrSection--2_G8xGa').items()
        items = doc('.Attrs--attr--33ShB6X').items()
        for item in items:
            product = item.text()
            count = 0
            for i in product:
                if i == "：":
                    break
                count += 1
            result[str(product[:count])] = str(product[count + 1:])
        print(result)

        time.sleep(1.5)
        comment_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div/div[2]/div[2]/div[2]/div[1]/div/div/div[2]/span')))  # 宝贝评价按钮
        comment_btn.click()
        time.sleep(1.5)
        html = driver.page_source
        doc = pq(html)
        for i in range(2):
            for y in range(10):
                js = 'window.scrollBy(0,450)'
                driver.execute_script(js)
                time.sleep(0.5)
            items = doc('.Comment--content--15w7fKj').items()  # 找到评论区
            for item in items:
                comment = item.text()
                # print(comment)
                comments.append(comment)
            try:
                # comment_btn = wait.until(EC.element_to_be_clickable(
                #     (By.XPATH, '//*[@id="root"]/div/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[3]/div/button[2]')))  # 下一页按钮
                # comment_btn.click()
                next_page_btn = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[3]/div/button[2]')
                driver.execute_script('arguments[0].click();', next_page_btn)
            except Exception as e:
                break
        print(comments)
        return str(result), comments, error

    except Exception as e:
        error = True
        return str(result), comments, error


def get_item(url):
    # url = 'https://item.taobao.com/item.htm?id=784281900551'
    if url[0:6] != 'https:':
        url = 'https:' + url
    print(url)
    driver.get(url)
    time.sleep(2)
    html = driver.page_source
    doc = pq(html)
    # items = doc('.ItemDetail--attrs--3t-mTb3').items()
    # items = doc('.Attrs--attrSection--2_G8xGa').items()
    items = doc('.Attrs--attr--33ShB6X').items()
    result = {}  # 手机参数
    for item in items:
        product = item.text()
        count = 0
        for i in product:
            if i == "：":
                break
            count += 1
        result[str(product[:count])] = str(product[count + 1:])
    time.sleep(1.5)
    comments = []  # 手机评论
    comment_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="root"]/div/div[2]/div[2]/div[2]/div[1]/div/div/div[2]/span')))  # 宝贝评价按钮
    comment_btn.click()
    time.sleep(2)
    html = driver.page_source
    doc = pq(html)
    items = doc('.Comment--content--15w7fKj').items()  # 找到评论
    for item in items:
        comment = item.text()
        comments.append(comment)
    return str(result), comments


def get_comment(url):  #
    # url = 'https://item.taobao.com/item.htm?id=782575688029'
    comments = []
    try:
        driver.get(url)
        time.sleep(2.5)
        comment_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div/div[2]/div[2]/div[2]/div[1]/div/div/div[2]/span')))  # 宝贝评价按钮
        comment_btn.click()
        time.sleep(1.5)
        html = driver.page_source
        doc = pq(html)
        # items = doc('.Comments--comments--1662-Lt').items()
        for i in range(1):
            items = doc('.Comment--content--15w7fKj').items()  # 找到评论区
            for item in items:
                comment = item.text()
                print(comment)
                comments.append(comment)
            comment_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="root"]/div/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[3]/div/button[2]')))  # 下一页按钮
            comment_btn.click()
        return comments
    except TimeoutException:
        print('TimeoutException')
        if len(comments):
            return comments
        else:
            print('No capture.')
            return 0


def get_tb_item():
    wb = openpyxl.load_workbook(filename="tb.xlsx")
    ws = wb['Sheet1']
    count = 1
    stop_time = 1
    continue_x = False
    i = 1
    workbook = xlsxwriter.Workbook('td_item.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, 'index')
    worksheet.write(0, 1, 'id')
    worksheet.write(0, 2, 'name')
    worksheet.write(0, 3, 'price')
    worksheet.write(0, 4, 'category')
    worksheet.write(0, 5, 'picture_addr')
    worksheet.write(0, 6, 'specifications')
    worksheet.write(0, 7, 'comment')
    wb_help = openpyxl.load_workbook(filename="td_item.xlsx")
    ws_help = wb_help['Sheet1']
    for row in ws_help.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, values_only=True):
        if i < count and continue_x:
            worksheet.write(i, 0, row[0])
            worksheet.write(i, 1, row[1])
            worksheet.write(i, 2, row[2])
            worksheet.write(i, 3, row[3])
            worksheet.write(i, 4, row[4])
            worksheet.write(i, 5, row[5])
            worksheet.write(i, 6, row[6])
            worksheet.write(i, 7, row[7])
            i += 1
        else:
            break
    i = 1
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, values_only=True):
        if i < count and continue_x:
            i += 1
            continue
        continue_x = False
        url = row[1]
        if url[0:6] != 'https:':
            url = 'https:' + url
        spec, comments, error = get_item_comment(url)
        if error:
            print(count)
            break
        worksheet.write(count, 0, row[0])
        worksheet.write(count, 1, url)
        worksheet.write(count, 2, row[2])
        worksheet.write(count, 3, row[3])
        worksheet.write(count, 4, row[4])
        worksheet.write(count, 5, row[5])
        worksheet.write(count, 6, spec)
        worksheet.write(count, 7, str(comments))
        # if len(comments):  # comments不为空
        #     for index, comment in enumerate(comments):
        #         worksheet.write(count, index + 7, comment)
        # else:
        #     print('empty')
        count += 1
        stop_time += 1
        if stop_time > 100:
            print(count)
            break
    workbook.close()


def read_only():
    wb = openpyxl.load_workbook(filename="td_item.xlsx")
    ws = wb['Sheet1']
    lis = []
    dic = {"id": None, "name": None, "price": None, "category": None, 'picture_addr': None, 'specification': None,
           'comments': None}
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, values_only=True):
        # print(row)
        print(row[0])
        if row[0] != None:
            dic["id"] = row[0]
            dic["name"] = row[1]
            dic["price"] = row[2]
            dic["category"] = row[3]
            dic["picture_addr"] = row[4]
            c = ast.literal_eval(row[5])
            # print(c, type(c))
            dic["specification"] = c
            dic['comments'] = row[6]
            lis.append(dic)
    # print(lis)
    return lis


if __name__ == '__main__':
    # is_loging = login_taobao()
    is_loging = True
    if is_loging:
        print('已经登录')
        time.sleep(3)
        # get_tb_item()  # 爬取手机参数和评价
        print(read_only())
