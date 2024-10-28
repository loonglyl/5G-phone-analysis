from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common import TimeoutException
from pyquery import PyQuery as pq
from urllib.parse import quote
import time
import openpyxl

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")  # 此处端口保持和命令行启动的端口一致
driver = Chrome(options=chrome_options, )  # executable_path='D:\Anaconda3\chrome-win64\chromedriver.exe'
wait = WebDriverWait(driver, 10)


wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0)

ws.cell(row=1, column=1, value="index")
ws.cell(row=1, column=2, value="id")
ws.cell(row=1, column=3, value="name")
ws.cell(row=1, column=4, value="price")
# 模拟淘宝登录
# chrome.exe --remote-debugging-port=9222 --user-data-dir='D:\chrome_data'


# 自动获取商品信息并自动翻页
def index_page():
    page, s, max_page = 1, 1, 3
    print("page=" + str(page) + ",s=" + str(s))
    try:
        for i in range(1, max_page):
            url = "https://search.jd.com/Search?keyword=5G手机&wq=5G手机&pvid=feb1e81188ca4a11964991bce3bbf228&isList=0&page=" + str(
                page) + "&s=" + str(s) + "&click=0"
            page = page + 1
            s = s + 30
            # res = requests.get(url, headers=headers)
            driver.get(url)
            time.sleep(1.5)
            # 获取商品信息
            html = driver.page_source
            doc = pq(html)

            items = doc('.gl-i-wrap').items()  # 所有手机商品
            count = 1
            for item in items:
                product = {'id': item.find('.p-icons').attr('id')[6:],
                           'price': item.find('.p-price')('i').text(),
                           'title': ''.join(item.find('.p-name')('a')('em').text().split('\n'))}
                count += 1
                ws.cell(row=count, column=1, value=str(count - 1))
                ws.cell(row=count, column=2, value=str(product['id']))
                ws.cell(row=count, column=3, value=str(product['title']))
                ws.cell(row=count, column=4, value=str(product['price']))
                print(product)
    except TimeoutException:
        print("Timeout")


if __name__ == '__main__':
    index_page()
    wb.save("jd.xlsx")  # 保存手机商品的网址信息
