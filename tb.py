from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common import TimeoutException
from pyquery import PyQuery as pq
from urllib.parse import quote
import time
import openpyxl

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")  # 此处端口保持和命令行启动的端口一致
driver = Chrome(options=chrome_options, )  # executable_path='D:\Anaconda3\chrome-win64\chromedriver.exe'
wait = WebDriverWait(driver, 10)


wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0)
count = 1
ws.cell(row=1, column=1, value="index")
ws.cell(row=1, column=2, value="id")
ws.cell(row=1, column=3, value="name")
ws.cell(row=1, column=4, value="price")
ws.cell(row=1, column=5, value="category")
ws.cell(row=1, column=6, value="picture_addr")


# 模拟淘宝登录
# chrome.exe --remote-debugging-port=9222 --user-data-dir='D:\chrome_data'
def login_taobao():
    login_url = 'https://login.taobao.com/member/login.jhtml'
    try:
        driver.get(login_url)
        input_login_id = wait.until(EC.presence_of_element_located((By.ID, 'fm-login-id')))
        input_login_password = wait.until(EC.presence_of_element_located((By.ID, 'fm-login-password')))
        input_login_id.send_keys('tb9792054524')  # 用你自己的淘宝账号替换  tb9792054524，tb64925158
        input_login_password.send_keys('Alihualin123')  # 用你自己的密码替换  Alihualin123，@zdm200212
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.fm-button.fm-submit.password-login')))
        submit.click()
        is_loging = wait.until(EC.url_changes(login_url))
        return is_loging
    except TimeoutException:
        print('login_taobao TimeoutException')
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.fm-button.fm-submit')))
        submit.click()
        is_loging = wait.until(EC.url_changes(login_url))
        if is_loging:
            return is_loging
        else:
            login_taobao()


# 解析获取商品信息
def get_products(category):
    """提取商品数据"""
    global count
    html = driver.page_source
    doc = pq(html)
    for y in range(28):
        js = 'window.scrollBy(0,200)'
        driver.execute_script(js)
        time.sleep(0.5)
    html = driver.page_source
    doc = pq(html)
    items = doc('.Card--doubleCardWrapper--L2XFE73').items()
    for item in items:
        product = {'url': item.attr('href'),
            'price': item.find('.Price--priceInt--ZlsSi_M').text(),
            'realsales': item.find('.Price--realSales--FhTZc7U').text(),
            'title': item.find('.Title--title--jCOPvpf').text(),
            'shop': item.find('.ShopInfo--TextAndPic--yH0AZfx').text(),
            'location': item.find('.Price--procity--_7Vt3mX').text(),
            'category':category,
            'picture_addr':item.find('.MainPic--mainPicWrapper--iv9Yv90')('img').attr('src')}
        count += 1
        if product['url'][0:6] != 'https:':
            product['url'] = 'https:' + product['url']
        ws.cell(row=count, column=1, value=str(count - 1))
        ws.cell(row=count, column=2, value=str(product['url']))
        ws.cell(row=count, column=3, value=str(product['title']))
        ws.cell(row=count, column=4, value=str(product['price']))
        ws.cell(row=count, column=5, value=str(product['category']))
        ws.cell(row=count, column=6, value=str(product['picture_addr']))
        print(product)


# 自动获取商品信息并自动翻页
def index_page(url, cur_page, max_page, catagory):
    try:
        driver.get(url)
        get_products(catagory)
        next_page_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//button/span[contains(text(),"下一页")]')))
        next_page_btn.click()
        do_change = wait.until(EC.url_changes(url))
        if do_change and cur_page < max_page:
            new_url = driver.current_url
            cur_page = cur_page + 1
            time.sleep(0.5)
            index_page(new_url, cur_page, max_page, catagory)
    except TimeoutException:
        print('---index_page TimeoutException---')


if __name__ == '__main__':
    # is_loging = login_taobao()
    category = ['华为', '小米', 'apple', '荣耀', '三星', '红米', 'oppo', 'vivo', '一加', '魅族', 'iqoo', '真我']
    is_loging = True
    if is_loging:
        print('已经登录')
        time.sleep(3)
        for i in category:
            KEYWORD = i + '手机'
            url = 'https://s.taobao.com/search?page=1&q=' + quote(KEYWORD) + '&tab=all'
            max_page = 3
            index_page(url, 1, max_page, i)
        wb.save("tb.xlsx")  # 保存手机商品的网址信息
