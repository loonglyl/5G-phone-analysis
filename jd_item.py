from lxml import etree
import requests
import time
import openpyxl
import json
import ast
import xlsxwriter


def read_only():
    wb = openpyxl.load_workbook(filename="jd_item.xlsx")
    ws = wb['Sheet1']
    lis = []
    dic = {"id": None, "name": None, "price": None, "specification": None}
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=5, max_row=ws.max_row, values_only=True):
        dic["id"] = row[0]
        dic["name"] = row[1]
        dic["price"] = row[2]
        print(row[3])
        c = ast.literal_eval(row[3])
        # print(c, type(c))
        dic["specification"] = c
        lis.append(dic)
    # print(lis)
    return lis


def get_jd_item():
    wb = openpyxl.load_workbook(filename="jd.xlsx")
    ws = wb['Sheet1']
    ws.cell(row=1, column=5, value="specifications")
    count = 1
    workbook = xlsxwriter.Workbook('jd_item.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, 'index')
    worksheet.write(0, 1, 'id')
    worksheet.write(0, 2, 'name')
    worksheet.write(0, 3, 'price')
    worksheet.write(0, 4, 'specifications')
    worksheet.write(0, 5, 'comments')
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, max_row=ws.max_row, values_only=True):
        worksheet.write(count, 0, row[0])
        worksheet.write(count, 1, row[1])
        worksheet.write(count, 2, row[2])
        worksheet.write(count, 3, row[3])
        worksheet.write(count, 4, jd_item(row[1]))
        count += 1
    workbook.close()


def jd_item(id):
    url = "https://item.jd.com/" + str(id) + ".html"   # https://item.jd.com/100097568128.html
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edge/124.0.0.0",
        }
    #  "Cookie": "shshshfpa=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; shshshfpx=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; __jdu=1712498482836958502465; pinId=BKmhqjK0Ot4FvWOdtn9TiA; pin=jd_lTmXtaLBvUsY; unick=jd_lTmXtaLBvUsY; _tp=7%2B9vkUgexJVrllJMczJ5lw%3D%3D; _pst=jd_lTmXtaLBvUsY; user-key=d4e0b5c4-efd8-4e69-87e6-a0f85f57fe3a; qrsc=3; unpl=JF8EALFnNSttWk1dAxsHThtHQ1oHW1gATEQBamQFAAlcTl1QSAodQUV7XlVdWBRKEh9vYRRVVVNKXA4aACsSEXteU11bD00VB2xXXAQDGhUQR09SWEBJJVlQXl4ITxcFZ2A1ZF5Ye1QEKwIdEhNOVV1bVAtCFQRnYwxQXFtOUwErAysVIHttU1ZYDUoQM25XBGQfDBdWAR0HGRtdS1tUXVgAQhIKbG4HU1VcQlAEGAccFiBKbVc; __jdv=76161171|haosou-search|t_262767352_haosousearch|cpc|5512151796_0_378714d8f96b4496b3421ee548db97bd|1713316144511; areaId=19; ipLoc-djd=19-1601-50258-129167; TrackID=1JSb7ijeU0EfdV5-iYXc4__5LdRAVrkfc97Iias3TkhIZsyRi6cA5FBrkNtYGRbP1WtiIvD7yahJxauK6S0JYVG9it7lQFlrvRA4QzTFYeSfUyEO84P7E9vWFMUezQeny; thor=676741B8CA3FCE6E194F2D84F06C5CB3344018AEFA0977C08C33E05EF2AEFFC2D0A42EA46D88CF2D6B178FDAAA985F63FAA583759B36AB62234AB1B33A45A455ACF514C7394AA90420946DB02EBFC9C9086C3CF93A55E3FBDCCB979E1BD583B396581BA939E938169295926E7780CF82164F7B763197BB10E8D1B3CC51D2C0BD2D5583BD01464691741F806AECA53B8CB703008F72C1F3F11C1FDB7DC4C60A0A; flash=2_Ts9f6zFn2Tqw0NqMxG4hqLTQ13GCG1ID25q2whi2C-wINYjIm1mJlK882w-DJEDo53i-11Mh_olckTVhMFBu14BGBMXX3jll3hQvglus2d5*; jsavif=1; jsavif=1; _gia_d=1; 3AB9D23F7A4B3CSS=jdd032ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQAAAAMPBW7FEQAAAAAACFT2A57E3VCIIUX; shshshfpb=BApXckia2DupAmC3SbfNMBEgpx2ei76hCBlAIg6169xJ1MrE1XoO2; avif=1; xapieid=jdd032ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQAAAAMPBW7FEQAAAAAACFT2A57E3VCIIUX; __jda=143920055.1712498482836958502465.1712498483.1713749255.1713922086.10; __jdb=143920055.5.1712498482836958502465|10.1713922086; __jdc=143920055; rkv=1.0; 3AB9D23F7A4B3C9B=2ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQ"
    res = requests.get(url, headers=headers)
    res.encoding = 'utf-8'
    text = res.text
    selector = etree.HTML(text)
    list = selector.xpath('//*[@class="Ptable"]')
    for i in list:
        title = i.xpath('.//div[@class="Ptable-item"]/h3/text()')
        content1 = i.xpath('.//div[@class="Ptable-item"]/dl/dl/dt/text()')
        content2 = i.xpath('.//div[@class="Ptable-item"]/dl/dl/dd/text()')
        if len(content1) != len(content2):
            count = 0
            for i in content1:
                if i == '入网型号':
                    break
                count += 1
            content2[count:count + 3] = []
        content = {}
        for i in range(len(content1)):
            content[str(content1[i])] = str(content2[i])
        result = str(content)
        print(result)
        return result


if __name__ == '__main__':
    # jd_item("100082427721")
    get_jd_item()
    # read_only()
