import ast

import openpyxl
import copy


def read_only(filename):
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb['Sheet1']
    lis = []
    link_lis = []
    dic = {"index": None, "id": None, "name": None, "price": None, "category": None, "picture_addr": None, "specifications": None, "feedback": None, "comments": None}
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, values_only=True):  # , max_col=5
        dic["index"] = row[0]
        temp = copy.deepcopy(row[1])
        link_lis.append(temp)
        temp = temp[20: -5]
        dic['id'] = temp
        dic['name'] = row[2]
        dic['price'] = row[3]
        dic["category"] = row[4]
        dic["picture_addr"] = row[5]
        c = ast.literal_eval(row[6])
        dic["specifications"] = c
        dic["feedback"] = row[7]
        dic["comments"] = row[8]
        lis.append(copy.deepcopy(dic))
    # print(lis)
    return lis, link_lis
