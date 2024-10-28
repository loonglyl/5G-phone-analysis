import numpy as np
import openpyxl
import ast
import matplotlib.pyplot as plt
import math
import os


def read_only(filepath):
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb['Sheet1']
    lis = []
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, values_only=True):
        if row[0] is not None:
            c = ast.literal_eval(row[5])
            lis.append(c)
    return lis


def plot_radar(labels1, values1, values2, filepath, number):
    """
    绘制双雷达图

    参数:
    labels1 (list): 第一组雷达图标签
    values1 (list): 第一组对应标签的数值
    values2 (list): 第二组对应标签的数值
    """
    # 确保两组标签和数值的长度相同
    assert len(labels1) == len(values1)
    assert len(labels1) == len(values2)
    for i in range(len(values1)):
        values1[i] = math.floor(values1[i] / 0.1) + 1  # +1是为了让图形比较自然，但层次范围变为了1~11
    for i in range(len(values2)):
        values2[i] = math.floor(values2[i] / 0.1) + 1
    # 计算雷达图的角度
    angles = [i / float(len(labels1)) * 2 * np.pi for i in range(len(labels1))]
    angles += angles[:1]  # 闭合图形
    # 支持中文
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    # 绘制雷达图
    fig = plt.figure(figsize=(8, 8))
    ax = fig.add_subplot(111, polar=True)

    # 设置标签
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)
    plt.xticks(angles[:-1], labels1)

    # 绘制第一组数值折线图
    ax.plot(angles, values1 + [values1[0]], linewidth=2, linestyle='solid', label='Current Phone')
    ax.fill(angles, values1 + [values1[0]], 'b', alpha=0.1)

    # 绘制第二组数值折线图
    ax.plot(angles, values2 + [values2[0]], linewidth=2, linestyle='solid', label='Average')
    ax.fill(angles, values2 + [values2[0]], 'r', alpha=0.1)

    # 设置图例
    ax.legend(loc='upper right', bbox_to_anchor=(1.1, 1.1))

    # 设置图表属性
    ax.set_rgrids(np.arange(1, max(max(values1), max(values2)) + 1, 1))
    ax.set_title("Radar Chart")
    ax.grid(True)
    if not os.path.exists("static/radar"):
        os.mkdir("static/radar")
    filename = "static/radar/" + str(number)
    plt.savefig(filename, dpi=75)


# filepath = 'tb_item.xlsx'
filepath = 'jd_item.xlsx'
number = 300


# CPU核心数，电池容量，前置摄像头像素，分辨率，存储容量，屏幕刷新率
def radar(filepath='jd_item.xlsx', number: int = 906):
    specification = read_only(filepath)
    parameter_list = np.zeros((len(specification), 6))
    search_keys = ['充电功率', '电池容量', '前摄主像素', '屏幕尺寸', '机身内存', '屏幕刷新率']
    # search_keys = ['CPU核心数', '电池容量', '前置摄像头像素', '分辨率', '存储容量', '屏幕刷新率']  # 淘宝5g手机参数列表
    mapping = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9,
               '十': 10, '十一': 11, '十二': 12, '十三': 13, '十四': 14, '十五': 15, '十六': 16,
               '百': 100, '千': 1000}
    for i in range(parameter_list.shape[0]):
        for j in range(parameter_list.shape[1]):
            if search_keys[j] in specification[i]:
                result = specification[i][search_keys[j]]
                if j == 0:  # 充电功率，~W
                    # if result.split('核')[0] in mapping.keys():
                    #     result = mapping[result.split('核')[0]]
                    # elif result.split('核')[0].isdigit():
                    #     result = int(result.split('核')[0])
                    # else:
                    #     result = 0
                    result = float(result.split('w')[0].split('W')[0])
                elif j == 1:  # 电池容量，~mAh(typ)
                    result = int(result.split('m')[0].split('M')[0])
                elif j == 2:  # 前摄主像素
                    result = result.split('万')[0]
                    total = 1
                    temp = ''
                    for letter in result:
                        if letter.isdigit():
                            temp += letter
                        else:  # 针对中文计量单位，有的多输入了一个"百"导致异常值
                            if letter in mapping.keys():
                                if int(temp) * mapping[letter] < 10000:
                                    total = int(temp) * mapping[letter]
                                else:  # 超出5000万像素的认为是异常值
                                    total = int(temp)
                                temp = ''
                                break
                    if temp == '':
                        result = total
                    else:
                        result = total * int(temp)
                elif j == 3:  # 屏幕尺寸，~英寸
                    result = float(result.split('英')[0])
                    # total = 1
                    # temp = ''
                    # for letter in result:
                    #     if letter.isdigit():
                    #         temp += letter
                    #     elif temp != '':
                    #         if total*int(temp) < 5e7:  # 大于5千万认为是异常值
                    #             total *= int(temp)
                    #         temp = ''
                    # if temp == '':
                    #     result = total
                    # else:  # temp不为空说明是数字
                    #     if total * int(temp) < 5e7:  # 大于5千万认为是异常值
                    #         result = total * int(temp)
                elif j == 4:  # 机身内存，~GB
                    # result = int(result.split('g')[0].split('G')[0])
                    total = 0
                    count = 1
                    if len(result.split(' ')) > 1:
                        count = len(result.split(' '))
                    elif len(result.split(',')) > 1:
                        count = len(result.split(','))
                    temp = ''
                    for letter in result:
                        if letter.isdigit():
                            temp += letter
                        elif temp != '':
                            if letter in ['T', 't']:  # 单位是TB，默认单位是GB
                                total += int(temp) * 1024
                            else:
                                total += int(temp)
                            temp = ''
                    if temp != '':
                        total += int(temp)
                    result = total // count  # 求平均值
                elif j == 5:  # 屏幕刷新率
                    if result.split('h')[0].split('H')[0][0].isdigit():  # 第一个字符是数字
                        result = int(result.split('h')[0].split('H')[0])
                    else:
                        result = 0
                parameter_list[i][j] = result
    min_values = np.array(np.min(parameter_list, axis=0), dtype=np.float32)
    max_values = np.array(np.max(parameter_list, axis=0), dtype=np.float32)
    # argmax = np.argmax(parameter_list, axis=0)
    parameter_list = (parameter_list - min_values) / (max_values - min_values)
    averages = np.average(parameter_list, axis=0)
    scale_table = np.zeros((10, 6), dtype=np.float32)
    for i in range(scale_table.shape[0]):
        scale_table[i] = max_values * (i + 1) / 10

    # 画比例表
    # plt.rcParams['font.sans-serif'] = ['SimHei']
    # plt.rcParams['axes.unicode_minus'] = False
    # fig = plt.figure(figsize=(12, 8))
    # ax = fig.add_subplot(111)
    # ax.axis('tight')
    # ax.axis('off')
    # search_keys = ['充电功率', '电池容量', '前摄主像素', '屏幕尺寸', '机身内存', '屏幕刷新率']
    # # columns = ['CPU核心数(个)', '电池容量(mAh)', '前置摄像头像素\n(万像素)', '分辨率(个像素)', '存储容量(GB)', '屏幕刷新率(Hz)']
    # columns = ['充电功率(W)', '电池容量(mAh)', '前摄主像素\n(万像素)', '屏幕尺寸(英寸)', '机身内存(GB)', '屏幕刷新率(Hz)']
    # table = ax.table(cellText=scale_table, rowLabels=[' ' + str(i + 2) + ' ' for i in range(10)], colLabels=columns,
    #                  loc='center', cellLoc='center', colColours=['yellow']*6, rowColours=['yellow']*10, bbox=[0, 0, 1, 1])
    # table.auto_set_font_size(False)
    # table.set_fontsize(12)
    # plt.show()

    # print(min_values)
    # print(max_values)
    # print(argmax)
    # for index in [number]:
    #     print(specification[index])
    #     print(parameter_list[index])
    # print(averages)
    plot_radar(search_keys, list(parameter_list[number]), list(averages), filepath, number)
